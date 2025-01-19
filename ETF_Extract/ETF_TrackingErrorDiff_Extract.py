#To extract the tracking error of the ETF 
#This code handles the trackingerror difference data
#Current process
#1.Download the data from the AMI website(Applying the filters)
#Future->Remove the manual process ,automate the whole flow
import openpyxl
import pandas as pd
import os
import re


# Function to fetch the value from 'Regular' or 'Direct' with fallback logic
def get_valid_value(row, regular_key, direct_key):
    # Check if 'Regular' column has a valid value
    regular_value = row.get(regular_key)
    if pd.notna(regular_value):  # Check for non-NaN values
        return regular_value
    # Fallback to 'Direct' column if 'Regular' is invalid
    return row.get(direct_key)

# Define the file path and file name
file_path = r"D:\\ETF_Data\\ETFProcessData\\Downloaded\\TrackingError"
file_name = "Dec_TrackingError.xlsx"
full_path = os.path.join(file_path, file_name)	


# Load the Excel file using openpyxl
wb = openpyxl.load_workbook(full_path)
sheet = wb.active

# Extract the month-year information from the first row
month_year_string = None
for row in sheet.iter_rows(values_only=True, max_row=3):  # Only look at the first->thrid row
    for cell in row:
        if cell and isinstance(cell, str) and 'Tracking Difference for' in cell:
            month_year_string = cell
            break
    if month_year_string:
        break

if not month_year_string:
    raise ValueError("Could not find 'Tracking Difference for [Month-Year]' string in the sheet.")

# Extract the month and year using regex
match = re.search(r'Tracking Difference for (\w+)-(\d{4})', month_year_string)
if not match:
    raise ValueError("Month and year information could not be extracted from the string.")
month, year = match.groups()



# Dynamically find the header row by searching for 'Tracking Difference (%)'
header_row = None
for i, row in enumerate(sheet.iter_rows(values_only=True), 1):
    if row and any('Tracking Difference (%)' in str(cell) for cell in row if cell is not None):
        header_row = i
        break

if header_row is None:
    raise ValueError("Header row with 'Tracking Difference (%)' not found.")

# Extract headers (merged cells) from the identified header rows
headers = []
for row in sheet.iter_rows(min_row=header_row, max_row=header_row + 2, values_only=True):  # Assuming three layers
    headers.append(row)

# Dynamically calculate rows to skip
rows_to_skip = header_row + len(headers)  # Skip header row and its subsequent layers

# Propagate headers for empty cells due to merged regions
propagated_headers = []
for row in headers:  # Iterate through each layer of headers
    propagated_row = []
    last_valid_value = None
    for value in row:
        if value is not None:  # Update the last valid value
            last_valid_value = value
        propagated_row.append(last_valid_value)
    propagated_headers.append(propagated_row)

 # Combine multi-row headers into single unique headers
final_headers = []
for col_idx in range(len(propagated_headers[0])):  # Iterate over columns
    combined_header = []
    for row in propagated_headers:  # Combine values across header layers
        combined_header.append(row[col_idx] if row[col_idx] else "")
    final_headers.append("_".join(map(str, combined_header)).strip("_"))


# Load the data into pandas, skipping rows dynamically
data = pd.read_excel(full_path, skiprows=rows_to_skip, header=None)



# Assign extracted headers to DataFrame
if len(final_headers) != len(data.columns):
    raise ValueError("Mismatch between extracted headers and DataFrame columns.")
data.columns = final_headers

# Add 'Month' and 'Year' columns to the DataFrame
data['Month'] = month
data['Year'] = year


# Filter data based on the given condition
filter_condition = (
    data['Scheme Name'].str.contains("ETF|Exchange Traded Fund", case=False, na=False) & 
    ~data['Scheme Name'].str.contains("FOF|Fund of Funds", case=False, na=False)
)

filtered_data = data[filter_condition]



# Create a list for processed rows
processed_rows = []
#Remove it after debugging
processed_df=[]

# Iterate through rows to pick values for 'Regular' or 'Direct'
for _, row in filtered_data.iterrows():
    # Check if the row contains the 'Scheme Name'
    if pd.notna(row.get('Scheme Name')):
        # Extract relevant data
        processed_row = {
        "Scheme Name": row.get("Scheme Name"),
            "1-Year": get_valid_value(row, "Tracking Difference (%)_1-Year_Regular", "Tracking Difference (%)_1-Year_Direct"),
            "3-Year": get_valid_value(row, "Tracking Difference (%)_3-Year_Regular", "Tracking Difference (%)_3-Year_Direct"),
            "5-Year": get_valid_value(row, "Tracking Difference (%)_5-Year_Regular", "Tracking Difference (%)_5-Year_Direct"),
            "10-Year": get_valid_value(row, "Tracking Difference (%)_10-Year_Regular", "Tracking Difference (%)_10-Year_Direct"),
            "Since Launch": get_valid_value(row, "Tracking Difference (%)_Since Launch_Regular", "Tracking Difference (%)_Since Launch_Direct"),
            "Month": month,
            "Year": year,
        }
        
        processed_rows.append(processed_row)
        processed_df.append(row)



# Convert the processed rows into a DataFrame
output_data = pd.DataFrame(processed_rows)

#Remove post debugging
output_df = pd.DataFrame(processed_df)

# Save the processed DataFrame to an Excel file
output_file_path = r"D:\\ETF_Data\\ETFProcessData\\Processed\\TrackingError"
output_file_name = f"{month}_{year}.xlsx"
output = os.path.join(output_file_path, output_file_name)
output_data.to_excel(output, index=False, sheet_name="Filtered Data")






