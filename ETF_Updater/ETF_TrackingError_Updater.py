#To update the Asset return details to the DB
import os
import re
import glob
import subprocess
import pymysql
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
import openpyxl
#To check/retrieve the ETF details
def get_etf_id(cur, scheme_name):
    
    # Primary lookup
    retrieve_sql = "SELECT `etf_id` FROM `etf` WHERE `etf_name` = UPPER(%s)"
    cur.execute(retrieve_sql, (scheme_name,))
    result = cur.fetchone()

    if result:
        return result[0]  # etf_id

    # Fallback to mapping table
    mapping_sql = "SELECT `etf_id` FROM `etf_mapping` WHERE `etf_name` = %s"
    cur.execute(mapping_sql, (scheme_name,))
    result = cur.fetchone()

    return result[0] if result else None


# DB connection setup
def connect_db():
    password = os.getenv('MYSQL_PASSWORD')
    connection = pymysql.connect(host='localhost',user='root',password=password,db='etf')
    return connection


# Load and normalize Benchmark Data into DataFrame
def load_tracking_difference_excel(path):
    # Load workbook and sheet
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # Step 1: Find the title row
    title_row = None
    title_text = None
    for row in ws.iter_rows(min_row=1, max_row=20):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('Tracking Difference for'):
                title_row = cell.row
                title_text = cell.value.strip()
                break
        if title_row:
            break

    if title_row is None:
        raise ValueError("Header starting with 'Tracking Difference for' not found.")

    print("Title row found at:", title_row)

    # Step 2: Dynamically find first non-empty row after title as header
    header_row = None
    for r in range(title_row + 1, title_row + 10):  # Look ahead max 10 rows
        values = [cell.value for cell in ws[r]]
        if any(v is not None for v in values):
            header_row = r
            break

    if header_row is None:
        raise ValueError("Could not find a valid header row after title.")

    print("Using header row:", header_row)

    # Step 3: Read Excel using found header row
    df = pd.read_excel(path, header=header_row - 1)
    print("Extracted columns:", df.columns.tolist())

    if 'scheme_name' not in df.columns:
        raise ValueError("Missing 'scheme_name' column in Excel.")

    # Step 4: Extract month info from the title row (not filename)
    match = re.search(r'Tracking Difference for (\w{3}-\d{4})', title_text)
    if match:
        monthinfo = match.group(1)
    else:
        raise ValueError("Could not extract month string like 'Mar-2025'.")

    return df, monthinfo

# Discovery: check which assets exist
def discovery_process(df, monthinfo, excluded_etfs):
    connection = connect_db(); 
    cursor = connection.cursor()
    missing_rows = []
    for _, row in df.iterrows():
        scheme_name = str(row["scheme_name"]).strip()
              
        #Skips the row in the list
        if scheme_name in excluded_etfs:
            print(f"Excluded ETF (ignored): {scheme_name}")
            continue  # Skip this row

        # ETF  lookup
        return_etfid=get_etf_id(cursor,scheme_name)

        if return_etfid:
            print(f"Found ETF: {scheme_name} -> ID: {return_etfid}")
        else:
            print(f"ETF not found: {scheme_name}")
            missing_rows.append(row)
    if missing_rows:
        out = pd.DataFrame(missing_rows)
        outfile = os.path.join(r'D:\ETF_Data\ETF_Error', f"MissingETF_DATA_{monthinfo}.xlsx")
        out.to_excel(outfile, index=False)
        print(f"Saved missing to {outfile}")
    cursor.close(); 
    connection.close()

# Update DB with returns
def update_db(df, monthinfo, excluded_assets):
    connection = connect_db()
    cursor = connection.cursor()
    

    insert_sql = (
        "INSERT INTO etf_trackingerror"
        " (etf_id, etf_trackingerror_timeperiod, etf_trackingerror_value, etf_trackingerror_month, etf_trackingerror_year)"
        " VALUES (%s, %s, %s, %s, %s)"
    )

    etftrackingerror_columns = {
        '1-Year (Regular)': '1Y',
        '1-Year (Direct)': '1Y',
        '3-Year (Regular)': '3Y',
        '3-Year (Direct)': '3Y',
        '5-Year (Regular)': '5Y',
        '5-Year (Direct)': '5Y',
        '10-Year (Regular)': '10Y',
        '10-Year (Direct)': '10Y',
        'Since-Launch (Regular)': 'SL',
        'Since-Launch (Direct)': 'SL'
    }

    # Step 1: Deduplicate by selecting the best row per asset
    best_rows = {}
    for _, row in df.iterrows():
        key = row['scheme_name']
        score = sum(not pd.isna(row.get(col)) for col in etftrackingerror_columns)  # Non-null count
        total = sum(float(row.get(col, 0)) if not pd.isna(row.get(col)) else 0 for col in etftrackingerror_columns)

        if key not in best_rows:
            best_rows[key] = (row, score, total)
        else:
            _, prev_score, prev_total = best_rows[key]
            if score > prev_score or (score == prev_score and total > prev_total):
                best_rows[key] = (row, score, total)

    # Step 2: Insert best row data into DB
    for etf_name, (row, _, _) in best_rows.items():
        if etf_name in excluded_assets:
            continue

        nav_date = pd.to_datetime(monthinfo)
        return_month = nav_date.strftime("%b")
        return_year = nav_date.year
        #Get the ETF id
        return_etfid=get_etf_id(cursor,etf_name)
      

        if return_etfid:
            etf_id = return_etfid
            for col, period_label in etftrackingerror_columns.items():
                value_raw = row.get(col)
                if pd.isna(value_raw):
                    print(f"Skipping {period_label} return for {etf_id} as it's missing.")
                    continue
                try:
                    value = float(value_raw)
                    #cursor.execute(insert_sql, (etf_id, period_label, value, return_month, return_year))
                    print("The DB values",etf_name,period_label,value)
                except Exception as e:
                    print(f"Failed to insert {period_label} return for {etf_id}: {e}")
        else:
            print(f"ETF not found: {etf_id}")

    connection.commit()
    cursor.close()
    connection.close()
           

if __name__ == '__main__':
    excluded_etfs = [
    "Bandhan BSE Sensex ETF",
    "SBI BSE SENSEX ETF",
    "Kotak BSE Sensex ETF",
    "Nippon India ETF BSE Sensex",
    "Nippon India ETF BSE Sensex Next 50",
    "SBI BSE 100 ETF",
    "SBI BSE Sensex ETF",
    "SBI BSE Sensex Next 50 ETF",
    "Motilal Oswal Nifty 50 ETF"

    ]
    loc = r'D:\ETF_Data\ETFDataProcessing\ETFProcessedData\TrackingError\ETF_Data_Mar-2025.xlsx'
    df, monthinfo = load_tracking_difference_excel(loc)
    discovery_process(df, monthinfo, excluded_etfs)
    update_db(df, monthinfo, excluded_etfs)
