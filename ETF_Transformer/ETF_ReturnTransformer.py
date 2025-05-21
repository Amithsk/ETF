import pandas as pd
import re
from openpyxl import load_workbook
from pathlib import Path

# Load workbook and sheet
excel_path = r'D:\ETF_Data\ETFDataProcessing\ETFRawData\AUM_Returns\Fund-Performance-28-Feb-2025.xlsx'
wb = load_workbook(excel_path, data_only=True)
ws = wb['Fund_Performance']

# Detect starting row based on 'Fund Performance' marker
start_row = None
for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
    if row and "Fund Performance" in str(row[0]):
        start_row = i
        break

if start_row is None:
    raise ValueError("Could not find 'Fund Performance' in the sheet.")

# Load DataFrame using pandas
df = pd.read_excel(excel_path, sheet_name='Fund_Performance', header=start_row)

# Clean column names
def clean_column_name(col: str) -> str:
    col = str(col).strip()
    col = col.replace('(%)', '')
    col = re.sub(r'[^a-zA-Z0-9]+', '_', col)
    col = re.sub(r'_+', '_', col)
    return col.strip('_').lower()

df.columns = [clean_column_name(col) for col in df.columns]

# Filter for ETFs, excluding FOFs
etf_df = df[
    df['scheme_name'].str.contains('ETF', case=False, na=False) &
    ~df['scheme_name'].str.contains(r'\bFOF\b|Fund of Fund', case=False, na=False)
].copy()

# Define time periods to extract
time_periods = ['1_year', '3_year', '5_year', '10_year', 'since_launch']

# Build base DataFrame (AUM dropped)
etf_final = pd.DataFrame()
etf_final['Scheme Name'] = etf_df['scheme_name']
etf_final['Benchmark'] = etf_df['benchmark']
etf_final['NAV Date'] = etf_df['nav_date']

# Add return columns with fallback logic
for period in time_periods:
    reg_col = f'return_{period}_regular'
    dir_col = f'return_{period}_direct'
    output_col = f'{period.replace("_", "-").title()} Return'

    if reg_col in etf_df.columns and dir_col in etf_df.columns:
        etf_final[output_col] = etf_df[reg_col].combine_first(etf_df[dir_col])
    elif reg_col in etf_df.columns:
        etf_final[output_col] = etf_df[reg_col]
    elif dir_col in etf_df.columns:
        etf_final[output_col] = etf_df[dir_col]
    else:
        etf_final[output_col] = pd.NA
        print(f"No return column found for {period}")

# Create month label
month_str = pd.to_datetime(etf_final['NAV Date'].iloc[0]).strftime('%b-%Y')
destination_dir = Path(r'D:\ETF_Data\ETFDataProcessing\ETFProcessedData\AUM_Returns')
destination_dir.mkdir(parents=True, exist_ok=True)

# Save ETF file
etf_file = destination_dir / f'ETF_Data_{month_str}.xlsx'
etf_final.to_excel(etf_file, index=False)

# Prepare Benchmark data
benchmark_columns = {
    'benchmark': 'Benchmark',
    'nav_date': 'Date',
    'return_1_year_benchmark': 'Return 1Year Benchmark',
    'return_3_year_benchmark': 'Return 3 Year Benchmark',
    'return_5_year_benchmark': 'Return 5 Year Benchmark',
    'return_10_year_benchmark': 'Return 10 Year Benchmark',
    'return_since_launch_benchmark': 'Return Since Launch Benchmark',
}

benchmark_fields = list(benchmark_columns.keys())
benchmark_rows = etf_df[benchmark_fields].dropna(subset=['benchmark'])

# Select best row per benchmark
best_rows = []
for benchmark, group in benchmark_rows.groupby('benchmark'):
    best_row = group.loc[group[benchmark_fields].notna().sum(axis=1).idxmax()]
    best_rows.append(best_row)

# Create final benchmark DataFrame
benchmark_df = pd.DataFrame(best_rows).rename(columns=benchmark_columns)

# Save benchmark file
benchmark_file = destination_dir / f'Benchmark_Data_{month_str}.xlsx'
benchmark_df.to_excel(benchmark_file, index=False)

# Save separate AUM file
etf_aum_df = etf_df[['scheme_name', 'nav_date', 'daily_aum_cr']].copy()
etf_aum_df['Month'] = pd.to_datetime(etf_aum_df['nav_date']).dt.strftime('%b-%Y')
etf_aum_df.rename(columns={
    'scheme_name': 'Scheme Name',
    'nav_date': 'NAV Date',
    'daily_aum_cr': 'AUM'
}, inplace=True)
etf_aum_df = etf_aum_df[['Scheme Name', 'Month', 'AUM']]

etf_aum_file = destination_dir / f'ETF_AUM_{month_str}.xlsx'
etf_aum_df.to_excel(etf_aum_file, index=False)

# Final confirmation
print("Files created:")
print(f"- ETF File: {etf_file}")
print(f"- Benchmark File: {benchmark_file}")
print(f"- AUM File: {etf_aum_file}")
