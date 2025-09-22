import pandas as pd
from pathlib import Path
import re
from openpyxl import load_workbook

def debug_etf_row(etf_name):
    matches = etf_df[etf_df['scheme_name'].str.contains(etf_name, case=False, na=False)]

    if matches.empty:
        print(f"\n No matching ETF found for: '{etf_name}'")
        return

    for idx, row in matches.iterrows():
        print(f"\n--- Debugging ETF: {row['scheme_name']} ---")
        print("\nFull row data:")
        print(row.to_string())

        print("\nTracking Difference Checks:")
        for period in time_periods:
            for variant in preferred_order:
                col_name = f"tracking_difference_{period}_{variant}"
                if col_name in etf_df.columns:
                    val = row.get(col_name)
                    status = f"{val}" if pd.notna(val) else " Missing"
                    print(f"  {col_name}: {status}")
                else:
                    print(f"Column not found: {col_name}")


# ===============================
# Configure paths (update if needed)
# ===============================
excel_path = r'D:\ETF_Data\ETFDataProcessing\ETFRawData\TrackingError\TrackingDifference_Aug_25.xlsx'
destination_dir = Path(r'D:\ETF_Data\ETFDataProcessing\ETFProcessedData\TrackingError')

# ===============================
# Detect format + build dataframe
# ===============================
wb = load_workbook(excel_path, data_only=True)
ws = wb.active

# Look for old-format title row "Tracking Difference ..." in first 20 rows
title_row = None
title_text = None
for row in ws.iter_rows(min_row=1, max_row=20):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and cell.value.strip().lower().startswith('tracking difference'):
            title_row = cell.row
            title_text = cell.value.strip()
            break
    if title_row:
        break

if title_row:
    # ---------------------------
    # Old multi-row header format
    # ---------------------------
    fmt = 'old_multi_header'
    # extract month from title text if possible (e.g., "Tracking Difference for Aug-2025")
    m = re.search(r'([A-Za-z]{3})[-_ ]?(\d{2,4})', title_text)
    if m:
        mon = m.group(1).title()
        yr = int(m.group(2))
        if yr < 100:
            yr += 2000
        month_str = f"{mon}-{yr}"
    else:
        month_str = 'UnknownMonth'

    # Read three header rows after the title and build combined headers
    header_1 = [str(cell.value).strip() if cell.value else '' for cell in ws[title_row + 1]]
    header_2 = [str(cell.value).strip() if cell.value else '' for cell in ws[title_row + 2]]
    header_3 = [str(cell.value).strip() if cell.value else '' for cell in ws[title_row + 3]]

    def clean_part(x):
        return re.sub(r'[^a-zA-Z0-9]+', '_', x.replace('(%)', '')).lower().strip('_')

    def forward_fill(lst):
        last = ''
        out = []
        for val in lst:
            if val:
                last = val
            out.append(last)
        return out

    header_1_filled = forward_fill(header_1)
    header_2_filled = forward_fill(header_2)
    header_3_filled = forward_fill(header_3)

    combined_headers = []
    for h1, h2, h3 in zip(header_1_filled, header_2_filled, header_3_filled):
        clean_h1 = clean_part(h1)
        clean_h2 = clean_part(h2)
        clean_h3 = clean_part(h3)

        # Special handling for scheme and benchmark
        if clean_h1.startswith("scheme_name") or clean_h1.startswith("benchmark"):
            combined_headers.append(clean_h1)
            continue

        if "since_launch" in clean_h2:
            full = f"tracking_difference_since_launch_{clean_h3}"
        else:
            parts = [clean_h1, clean_h2, clean_h3]
            full = "_".join(p for p in parts if p)

        combined_headers.append(full.strip('_'))

    # Read data rows below the 3 header rows
    # skiprows should equal the number of rows before the actual data (title_row + 3 header rows)
    skiprows = title_row + 3
    df = pd.read_excel(excel_path, header=None, skiprows=skiprows)

    # Align header length to number of columns
    ncols = df.shape[1]
    if len(combined_headers) < ncols:
        for i in range(len(combined_headers), ncols):
            combined_headers.append(f'col_{i}')
    elif len(combined_headers) > ncols:
        combined_headers = combined_headers[:ncols]

    df.columns = combined_headers

else:
    # ---------------------------
    # New single-row header format
    # ---------------------------
    fmt = 'new_single_header'
    # Extract month from filename (e.g., TrackingDifference_Aug_25.xlsx)
    stem = Path(excel_path).stem
    m = re.search(r'([A-Za-z]{3})[-_ ]?(\d{2,4})', stem)
    if m:
        mon = m.group(1).title()
        yr = int(m.group(2))
        if yr < 100:
            yr += 2000
        month_str = f"{mon}-{yr}"
    else:
        month_str = 'UnknownMonth'

    # Load dataframe with header row 0 and normalize column names
    df = pd.read_excel(excel_path, header=0)
    df.columns = [re.sub(r'[^a-zA-Z0-9]+', '_', str(c)).lower().strip('_') for c in df.columns]

    # rename period columns to include expected prefix "tracking_difference_..."
    col_renames = {}
    for col in df.columns:
        # matches '1_year_regular', '3_year_direct', etc.
        m1 = re.match(r'^(\d+)_year_(regular|direct)$', col)
        m2 = re.match(r'^(since_launch)_(regular|direct)$', col)
        if m1:
            col_renames[col] = f"tracking_difference_{m1.group(1)}_year_{m1.group(2)}"
        elif m2:
            col_renames[col] = f"tracking_difference_{m2.group(1)}_{m2.group(2)}"
    if col_renames:
        df = df.rename(columns=col_renames)

# Drop fully empty rows
df = df.dropna(how='all')

# Ensure scheme_name and benchmark exist (normalize if necessary)
if 'scheme_name' not in df.columns:
    for c in df.columns:
        if 'scheme' in c:
            df = df.rename(columns={c: 'scheme_name'})
            break
if 'benchmark' not in df.columns:
    for c in df.columns:
        if 'benchmark' in c:
            df = df.rename(columns={c: 'benchmark'})
            break

# ------------------------------
# Filter ETF rows
# ------------------------------
etf_df = df[
    df['scheme_name'].str.contains('ETF', case=False, na=False) &
    ~df['scheme_name'].str.contains(r'\bFOF\b|Fund of Fund', case=False, na=False)
    ].copy()

# These are your preferred column suffixes
time_periods = ['1_year', '3_year', '5_year', '10_year', 'since_launch']
preferred_order = ['regular', 'direct']  # prefer regular, fallback to direct

# Always include scheme_name and benchmark
final_columns = {'scheme_name': 'scheme_name', 'benchmark': 'benchmark'}

# Dynamically choose the first available column with actual data (not NaN)
for period in time_periods:
    selected = False
    for variant in preferred_order:
        col_name = f"tracking_difference_{period}_{variant}"
        if col_name in etf_df.columns and etf_df[col_name].notna().sum() > 0:
            label = period.replace('_', '-').title()
            final_columns[col_name] = f"{label} ({variant.title()})"
            selected = True
            break
    # fallback: sometimes the prefix wasn't added; accept columns like "1_year_regular"
    if not selected:
        for variant in preferred_order:
            alt = f"{period}_{variant}"
            if alt in etf_df.columns and etf_df[alt].notna().sum() > 0:
                final_columns[alt] = f"{period.replace('_', '-').title()} ({variant.title()})"
                selected = True
                break
    if not selected:
        print(f"No usable tracking difference column for {period}")

# Select only available columns to avoid KeyError
available_cols = [c for c in final_columns.keys() if c in etf_df.columns]
if 'scheme_name' not in available_cols:
    raise ValueError("Could not find 'scheme_name' column after normalization.")

etf_rows = etf_df[available_cols].dropna(subset=['scheme_name'])
etf_final = etf_rows.rename(columns=final_columns)

# Save output
etf_file = destination_dir / f'ETF_Data_{month_str}.xlsx'
etf_final.to_excel(etf_file, index=False, startrow=1)

# Add the month string at the top
wb_out = load_workbook(etf_file)
ws_out = wb_out.active
ws_out.insert_rows(1)
ws_out['A1'] = f"Tracking Difference for {month_str}"
wb_out.save(etf_file)

print(f"Format detected: {fmt}")
print(f"Month extracted: {month_str}")
print(f"Saved: {etf_file}")
